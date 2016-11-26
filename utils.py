# coding=gbk
'''
Created on 2016��2��5��

@author: ����
'''
import os, sys

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import datetime


def getHome():
    p = sys.path[0]
    if os.path.isdir(p):
        return p
    elif os.path.isfile(p):
        return os.path.dirname(p)

'''
just convert the first sheet to csv
'''
def Excel2CSV(ExcelFile, CSVFile):
    department = os.path.basename(ExcelFile).split(".")[0]
    isDepartment_column = False
    workbook = load_workbook(ExcelFile, read_only=True)
    worksheet = workbook.active
    rowidx = 0
    with open(CSVFile, 'wb') as csvfile:
        for row in worksheet.rows:
            linevalue = []
            for cell in row:
                #print(cell.data_type)
                if cell.data_type == Cell.TYPE_NULL:
                    linevalue.append(str(cell.value))
                    #print('error: ' + str(cell.value))
                else:
                    b = cell.value
                    if type(b) == type(4.0):
                        b= str(b)
                    linevalue.append(b)

            #���벿����Ϣ
            #
            #��һ�в���head���ڶ��п�ʼ�����ļ���-��������        
            if rowidx == 0:
                #��ȡ"����"�����Ƿ����
                try:
                    if linevalue.index("����", ) > -1:
                        isDepartment_column = True
                except Exception:
                    isDepartment_column = False
                    linevalue.append("����")
            else:
                if not isDepartment_column:
                    linevalue.append(department)
                
            linevalue.append("\n")
            d = ','.join(linevalue).encode('utf-8')
            csvfile.write(d)
            rowidx = rowidx + 1
            

def writeExcel(data, file, startrow, startcol):
    workbook  = load_workbook(getHome() + "/template/template.xlsx")
    worksheet = workbook.active
    rowidx = 0
    for row in data:
        colidx = 0
        for col in row:
            worksheet.cell(row=startrow+rowidx, column=startcol+colidx, value=col)
            colidx = colidx + 1
        rowidx = rowidx + 1    
    workbook.save(file)
    
def writeCreateTime(file, startrow, startcol):
    workbook  = load_workbook(file)
    worksheet = workbook.active
    worksheet.cell(row=startrow, column=startcol, value=datetime.datetime.now())        
    workbook.save(file)